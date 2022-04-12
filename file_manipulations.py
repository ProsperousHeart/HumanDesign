# imports
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, Font
# from openpyxl.styles.colors import BLUE

dest_filename = "Human Design Data.xlsx"

perm_err = """\nUnable to save file - it is likely open.
Please close the file and press ENTER to continue ...
(If the file is already closed, this program will error out...)\n"""

# read in a file - allow them to choose a file

# export to XLSX
def create_workbook(info2prop: dict):
    """
    This function takes in a file name string and sets up
    the empty workbook with which to add data to.
    It returns the Workbook object to manipulate the data.
    """

    def create_style(wb):
        """

        :param wb:
        :return:
        """

        bold = NamedStyle(name="bold")
        bold.font = Font(bold=True)
        wb.add_named_style(bold)

    def create_sheets(new_wb: Workbook, sheets: list, data_dict: dict):
        """

        :param new_wb:
        :param sheets:
        :param data_dict:
        :return:
        """

        for idx0, sheet in enumerate(sheets):
            if idx0 == 0:
                wbs = new_wb.active
                wbs.title = sheet
            else:
                wbs = new_wb.create_sheet(sheet)

            cols = list()
            if sheet == "Type":
                cols = ["HD Type", "Energy Type", "Strategy", "Signature", "Not-Self", "Pages"]
            elif sheet == "Centers":
                cols = ["Center", "Explanations", "Pages", "Definition", "Gates"]

            if len(cols) == 0:
                print("There was an issue with creating sheets!")
            else:
                dct_list = data_dict[sheet]
                for idx, col_name in enumerate(cols):
                    wbs.cell(row=1, column=idx+1).value = col_name
                    for idx2, item in enumerate(dct_list):
                        idx2 += 1
                        wbs.cell(row=1+idx2, column=idx+1).value = str(item[col_name])

    workbook = Workbook()
    create_style(workbook)
    sheet_list = ["Type", "Centers"]
    create_sheets(workbook, sheet_list, info2prop)
    return workbook


def write2file(file_type: int, data_lst: dict):
    """

    :param file_type:
    :param data_lst:
    :return:
    """

    if file_type == 1:  # XLSX file
        wb = create_workbook(data_lst)
        try:
            wb.save(dest_filename)
        except PermissionError as err:
            input(perm_err)
        finally:
            try:
                wb.save(dest_filename)
            except PermissionError as err:
                print("Unfortunately there is another permissions issue. Exiting the program.")
    else:
        print("Other file types (like exporting to Google sheets) are not yet written.")

# export to Google sheet
